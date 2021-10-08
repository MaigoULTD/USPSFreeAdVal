from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import time
import openpyxl



options = webdriver.ChromeOptions()
options.add_argument('--ignore-certificate-errors-spki-list')
options.add_argument('--ignore-ssl-errors')
options.add_argument('--ignore-certficate-errors')

driver = webdriver.Chrome(executable_path=r'CHROMEDRIVER_DIR', options=options) #Init driver
driver.get("https://tools.usps.com/zip-code-lookup.htm?byaddress") #Get USPS Page


workbook = openpyxl.load_workbook('WORKBOOK')#Load excel sheet
#address_sheet = workbook['addresses']#Load sheet for addresses without apts or units.

address_sheet = workbook['WORKSHEET'] #Load sheet for apartments. Uncomment to use with script.
saveindex=0
index=44
while index != 64:    
    ## Need to fix this part
    tAddress = ((address_sheet.cell(row=index,column=6)).value) #Get address at row
    tApt = ((address_sheet.cell(row=index,column=7)).value)#Get apartments. Uncomment to use apt sheet
    tCity = ((address_sheet.cell(row=index,column=8)).value) #Get city at row
    tZipbyaddress = ((address_sheet.cell(row=index,column=10)).value) #Get Zipcode
    ##


    driver.find_element_by_id("tAddress").send_keys(tAddress)
    driver.find_element_by_id("tCity").send_keys(tCity)
    driver.find_element_by_id("tApt").send_keys(tApt)
    # driver.find_element_by_id("tState").send_keys("apt") # Dropdown menu. Might not need it if Zipcode is legitimate. Would help with verification. 
    driver.find_element_by_id("tZip-byaddress").send_keys(tZipbyaddress)
    driver.find_element_by_id("zip-by-address").click() # Shitty ID name. LMAO

    #Because the Address List page is dynamically loaded, trying to wait until Selenium can located the zipcode-result-address section, which loads if an address is found. Timelimit is 10 seconds.
    try:
        element = WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.CLASS_NAME,"zipcode-result-address")))
    except (TimeoutException):
        address_sheet.cell(row=index, column=18).value='UNVERIFIED' # Selenium throws a TimeoutException if no element is found in the specified time. Excepting this will let python enter values into the excel sheet. 
    else:
        address_sheet.cell(row=index, column=18).value='Verified'
    
    index=index+1
    driver.refresh()
    #optional. change depending on how frequent you want the script to save. Default is every 20.
    if saveindex == 20:
        workbook.save('WORKBOOK')
        saveindex = 0
    else:
        saveindex=saveindex+1
        continue

workbook.save('WORKBOOK')
driver.close()
workbook.close()